import os
import json
import openai
import pandas as pd
from dotenv import load_dotenv
from tqdm import tqdm
from openpyxl import load_workbook
from datetime import datetime
import glob

load_dotenv()
openai.api_key = os.getenv("OPENAI_API_KEY")
client = openai

def get_fire_related_score(content):
    prompt = (
        "On a scale of 0 to 10, how strongly is the following tweet related to fire damages or destruction in the United States? "
        "A score of 0 means not related at all, 10 means it is definitely about fire damages or destruction in the USA. "
        "Only use the tweet content for your evaluation.\n\n"
        f"Tweet content: {content[:2000]}"
    )
    messages = [
        {"role": "system", "content": "You are an AI that rates the fire-relatedness of tweets about fire damages or destruction in the USA. Respond with a single integer from 0 to 10."},
        {"role": "user", "content": prompt}
    ]
    try:
        ai_response = client.chat.completions.create(
            model='gpt-4o-mini',
            messages=messages,
            temperature=0,
        )
        answer = ai_response.choices[0].message.content.strip()
        import re
        match = re.search(r'\b(10|[0-9])\b', answer)
        if match:
            return int(match.group(1))
        return answer
    except Exception as e:
        print(f"Error with OpenAI API (score): {e}")
        return ""

def verify_fire_incident(title, content, url, country="USA"):
    print(url)
    truncated_content = content[:4000]
    fire_incident_prompt = (
        "You are given the content of a tweet or news snippet. Determine if it describes a fire incident in the United States that likely caused damage to physical structures (such as homes, apartments, offices, commercial buildings, factories, or infrastructure). "
        "The fire may have resulted in structural damage or destruction, due to causes like electrical faults, negligence, accidents, natural disasters (e.g., wildfires), or arson. "
        "Be inclusive: If the tweet/news suggests a fire incident with possible or likely damage to structures, even if not 100% explicit, respond with 'yes'. "
        "Respond with 'yes' if the tweet/news is about a fire incident in the USA that could have caused damage to physical structures. Otherwise, respond with 'no'.\n\n"
        f"Content: {truncated_content}\nURL: {url}\n"
        "Only use the provided content for your evaluation. Do not infer or assume details not present in the text, but err on the side of inclusion if the fire incident is plausible."
    )
    messages = [
        {
            "role": "system",
            "content": "You are an AI tasked with evaluating tweets to determine if they describe fire damages or destruction in the United States. Be inclusive: If the tweet/news is plausibly about fire damages or destruction in the USA, mark as 'yes'."
        },
        {"role": "user", "content": fire_incident_prompt}
    ]
    try:
        ai_response = client.chat.completions.create(
            model='gpt-4o-mini',
            messages=messages,
            temperature=0,
        )
        answer = ai_response.choices[0].message.content.strip()
        print(answer)
        return answer
    except Exception as e:
        print(f"Error with OpenAI API: {e}")
        return "no"

def update_live_json(live_json_path, entry):
    import threading
    lock = threading.Lock()
    lock.acquire()
    try:
        if os.path.exists(live_json_path):
            with open(live_json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        else:
            data = []
        data.append(entry)
        with open(live_json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    finally:
        lock.release()

def autosize_and_format_excel(excel_path):
    from openpyxl.utils import get_column_letter
    wb = load_workbook(excel_path)
    ws = wb.active
    default_width = 30
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = default_width
        for cell in col:
            cell.alignment = cell.alignment.copy(wrap_text=True)
    for row in ws.iter_rows():
        max_height = 15
        for cell in row:
            if cell.value:
                lines = str(cell.value).count("\n") + 1
                length = len(str(cell.value))
                est_height = max(15, min(150, lines * 15 + length // 50 * 15))
                if est_height > max_height:
                    max_height = est_height
        ws.row_dimensions[row[0].row].height = max_height
    url_col = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value and str(cell.value).lower() == "url":
            url_col = idx
            break
    if url_col:
        for row in ws.iter_rows(min_row=2, min_col=url_col, max_col=url_col):
            for cell in row:
                if cell.value and str(cell.value).startswith("http"):
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"
    wb.save(excel_path)

def verify_and_save_to_excel(cleaned_json_path, excel_path=None, live_json_path=None):
    import os
    import pandas as pd
    # Generate timestamped filenames
    dt_str = datetime.now().strftime('%d%b_%H%M').lower()
    if excel_path is None:
        excel_path = f"output/verified_fires_{dt_str}.xlsx"
    if live_json_path is None:
        live_json_path = f"output/live_verified_fires_{dt_str}.json"
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    os.makedirs(os.path.dirname(live_json_path), exist_ok=True)
    with open(cleaned_json_path, "r", encoding="utf-8") as f:
        tweets = json.load(f)
    verified_rows = []
    for tweet in tqdm(tweets, desc="Verifying tweets with AI"):
        title = tweet.get("content", "")[:100]
        content = tweet.get("content", "")
        date = tweet.get("timestamp", "")
        url = tweet.get("tweet_url", "")
        source = tweet.get("username", "")
        verification_result = verify_fire_incident(title, content, url, country="USA")
        fire_related_score = get_fire_related_score(content) if verification_result.lower().startswith("yes") else 0
        verified_at = datetime.now().isoformat()
        print(f"{date} {verification_result.strip().lower()} {url}")
        if verification_result.lower().startswith("yes"):
            row = {
                "title": title,
                "content": content,
                "published_date": date,
                "url": url,
                "source": source,
                "fire_related_score": fire_related_score,
                "verification_result": verification_result,
                "verified_at": verified_at
            }
            verified_rows.append(row)
            update_live_json(live_json_path, row)
            if os.path.exists(excel_path):
                df_existing = pd.read_excel(excel_path)
                df_new = pd.DataFrame([row])
                df_combined = pd.concat([df_existing, df_new], ignore_index=True)
                df_combined.to_excel(excel_path, index=False)
            else:
                df = pd.DataFrame([row])
                df.to_excel(excel_path, index=False)
            autosize_and_format_excel(excel_path)
    if verified_rows:
        print(f"✅ Saved {len(verified_rows)} verified fire incidents to {excel_path} and {live_json_path}")
        import smtplib
        from email.message import EmailMessage
        EMAIL_HOST = 'smtp.gmail.com'
        EMAIL_PORT = 587
        EMAIL_HOST_USER = 'agilemorphsolutions@gmail.com'
        EMAIL_HOST_PASSWORD = 'vktnzpaaurneigpg'
        TO_EMAIL = 'forrohitsingh99@gmail.com'
        def send_email_with_attachment(subject, body, to_email, attachment_paths):
            msg = EmailMessage()
            msg['Subject'] = subject
            msg['From'] = EMAIL_HOST_USER
            msg['To'] = to_email
            msg.set_content(body)
            for attachment_path in attachment_paths:
                with open(attachment_path, 'rb') as f:
                    file_data = f.read()
                    file_name = os.path.basename(attachment_path)
                msg.add_attachment(file_data, maintype='application', subtype='octet-stream', filename=file_name)
            with smtplib.SMTP(EMAIL_HOST, EMAIL_PORT) as server:
                server.starttls()
                server.login(EMAIL_HOST_USER, EMAIL_HOST_PASSWORD)
                server.send_message(msg)
        send_email_with_attachment(
            subject="Verified Fire Incidents - Latest Batch",
            body="Please find attached the latest verified fire incidents (Excel and JSON).",
            to_email=TO_EMAIL,
            attachment_paths=[excel_path, live_json_path]
        )
        print(f"📧 Email sent to {TO_EMAIL} with attachments {excel_path} and {live_json_path}")
    else:
        print("No verified fire incidents found.")

if __name__ == "__main__":
    import sys
    dt_str = datetime.now().strftime('%d%b_%H%M').lower()
    if len(sys.argv) > 1:
        json_path = sys.argv[1]
    else:
        filtered_files = sorted(glob.glob("*_cleaned_tweets.json"), reverse=True)
        if filtered_files:
            json_path = filtered_files[0]
            print(f"No input file specified. Using latest cleaned file: {json_path}")
        else:
            print("No cleaned tweets file found. Please run tweet_fire_search.py first or specify a file.")
            exit(1)
    excel_path = f"output/verified_fires_{dt_str}.xlsx"
    live_json_path = f"output/live_verified_fires_{dt_str}.json"
    verify_and_save_to_excel(json_path, excel_path, live_json_path) 