import os
import json
import openai
import pandas as pd
from dotenv import load_dotenv
from tqdm import tqdm
from openpyxl import load_workbook
from datetime import datetime, timedelta
import re

load_dotenv()
openai.api_key = os.getenv("OPENAI_API_KEY")
client = openai

FIRE_INCIDENT_KEYWORDS = [
    "burn", "evacuate", "evacuation", "damage", "destroy", "blaze", "smoke", "flames", "emergency", "brushfire", "structure fire", "forest fire", "house fire", "apartment fire", "building fire", "outbreak", "spread"
]
STRUCTURE_DAMAGE_KEYWORDS = [
    "structure fire", "building fire", "house fire", "apartment fire", "commercial fire", "warehouse fire", "residential fire", "industrial fire", "office fire", "school fire", "church fire", "hospital fire", "barn fire", "garage fire", "hotel fire", "motel fire", "condo fire", "duplex fire", "multi-family fire", "business fire", "restaurant fire", "store fire", "shopping center fire", "mall fire",
    "destroyed", "damaged", "total loss", "collapsed", "evacuated"
]
US_LOCATIONS = [
    "Alabama", "Alaska", "Arizona", "Arkansas", "California", "Colorado", "Connecticut", "Delaware", "Florida",
    "Georgia", "Hawaii", "Idaho", "Illinois", "Indiana", "Iowa", "Kansas", "Kentucky", "Louisiana", "Maine",
    "Maryland", "Massachusetts", "Michigan", "Minnesota", "Mississippi", "Missouri", "Montana", "Nebraska",
    "Nevada", "New Hampshire", "New Jersey", "New Mexico", "New York", "North Carolina", "North Dakota", "Ohio",
    "Oklahoma", "Oregon", "Pennsylvania", "Rhode Island", "South Carolina", "South Dakota", "Tennessee", "Texas",
    "Utah", "Vermont", "Virginia", "Washington", "West Virginia", "Wisconsin", "Wyoming",
    "New York", "Los Angeles", "Chicago", "Houston", "Phoenix", "Philadelphia", "San Antonio", "San Diego", "Dallas",
    "San Jose", "Austin", "Jacksonville", "Fort Worth", "Columbus", "Charlotte", "San Francisco", "Indianapolis",
    "Seattle", "Denver", "Washington", "Boston", "El Paso", "Nashville", "Detroit", "Oklahoma City", "Portland",
    "Las Vegas", "Memphis", "Louisville", "Baltimore", "Milwaukee", "Albuquerque", "Tucson", "Fresno", "Sacramento",
    "Mesa", "Kansas City", "Atlanta", "Omaha", "Colorado Springs", "Raleigh", "Miami", "Long Beach", "Virginia Beach",
    "Oakland", "Minneapolis", "Tulsa", "Tampa", "Arlington"
]

def is_within_last_72_hours(iso_timestamp):
    try:
        tweet_time = datetime.fromisoformat(iso_timestamp.replace('Z', '+00:00'))
        now = datetime.now(tweet_time.tzinfo)
        return (now - tweet_time).total_seconds() <= 72 * 3600
    except Exception:
        return False

def is_relevant_tweet(tweet):
    content = tweet.get("content", "").lower()
    fire_present = any(kw in content for kw in FIRE_INCIDENT_KEYWORDS)
    structure_damage_present = any(kw in content for kw in STRUCTURE_DAMAGE_KEYWORDS)
    location_present = any(loc.lower() in content for loc in US_LOCATIONS)
    return fire_present and (structure_damage_present or location_present)

def clean_tweets_json(raw_path, cleaned_path):
    with open(raw_path, "r", encoding="utf-8") as f:
        tweets = json.load(f)
    cleaned = [tw for tw in tweets if is_within_last_72_hours(tw.get("timestamp", ""))]
    with open(cleaned_path, "w", encoding="utf-8") as f:
        json.dump(cleaned, f, indent=4, ensure_ascii=False)
    print(f"🧹 Cleaned {cleaned_path}: {len(cleaned)} recent tweets remain.")

def verify_fire_incident(title, content, url, country="USA"):
    print(url)
    truncated_content = content[:4000]
    fire_incident_prompt = (
    "You are given the content of a tweet or news snippet. Determine if it describes a fire incident in the United States that likely caused damage to physical structures (such as homes, apartments, offices, commercial buildings, factories, or infrastructure). "
    "The fire may have resulted in structural damage or destruction, due to causes like electrical faults, negligence, accidents, natural disasters (e.g., wildfires), or arson. "
    "Be inclusive: If the tweet/news suggests a fire incident with possible or likely damage to structures, even if not 100% explicit, respond with 'yes'. "
    "Respond with 'yes' if the tweet/news is about a fire incident in the USA that could have caused damage to physical structures. Otherwise, respond with 'no'.\n\n"
    f"Content: {truncated_content}\nURL: {url}\n"
    "Only use the provided content for your evaluation. Do not infer or assume details not present in the text, but err on the side of inclusion if the fire incident is plausible."
)
    messages = [
        {
            "role": "system",
            "content": "You are an AI tasked with evaluating tweets to determine if they describe fire damages or destruction in the United States. Be inclusive: If the tweet/news is plausibly about fire damages or destruction in the USA, mark as 'yes'."
        },
        {"role": "user", "content": fire_incident_prompt}
    ]
    try:
        ai_response = client.chat.completions.create(
            model='gpt-4o-mini',
            messages=messages,
            temperature=0,
        )
        answer = ai_response.choices[0].message.content.strip()
        print(answer)
        return answer
    except Exception as e:
        print(f"Error with OpenAI API: {e}")
        return "no"

def get_fire_related_score(content):
    prompt = (
        "On a scale of 0 to 10, how strongly is the following tweet related to fire damages or destruction in the United States? "
        "A score of 0 means not related at all, 10 means it is definitely about fire damages or destruction in the USA. "
        "Only use the tweet content for your evaluation.\n\n"
        f"Tweet content: {content[:2000]}"
    )
    messages = [
        {"role": "system", "content": "You are an AI that rates the fire-relatedness of tweets about fire damages or destruction in the USA. Respond with a single integer from 0 to 10."},
        {"role": "user", "content": prompt}
    ]
    try:
        ai_response = client.chat.completions.create(
            model='gpt-4o-mini',
            messages=messages,
            temperature=0,
        )
        answer = ai_response.choices[0].message.content.strip()
        # Extract the first integer in the answer
        match = re.search(r'\b(10|[0-9])\b', answer)
        if match:
            return int(match.group(1))
        return answer
    except Exception as e:
        print(f"Error with OpenAI API (score): {e}")
        return ""

def update_live_json(live_json_path, row):
    if os.path.exists(live_json_path):
        with open(live_json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        data.append(row)
        with open(live_json_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4)
    else:
        with open(live_json_path, "w", encoding="utf-8") as f:
            json.dump([row], f, indent=4)

def autosize_and_format_excel(excel_path):
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                cell.number_format = '@' # Format as text
        for col in ws.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2 # Adjust as needed
            ws.column_dimensions[column].width = adjusted_width
        wb.save(excel_path)
    except Exception as e:
        print(f"Error autosizing Excel: {e}")

def verify_and_save_to_excel(cleaned_json_path, excel_path="output/verified_fires.xlsx", live_json_path="output/live_verified_fires.json"):
    import os
    import pandas as pd
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    os.makedirs(os.path.dirname(live_json_path), exist_ok=True)
    with open(cleaned_json_path, "r", encoding="utf-8") as f:
        tweets = json.load(f)
    # Filter for tweets within the last 72 hours
    tweets = [tw for tw in tweets if is_within_last_72_hours(tw.get("timestamp", ""))]
    verified_rows = []
    for tweet in tqdm(tweets, desc="Verifying tweets with AI"):
        title = tweet.get("content", "")[:100]
        content = tweet.get("content", "")
        date = tweet.get("timestamp", "")
        url = tweet.get("tweet_url", "")
        source = tweet.get("username", "")
        verification_result = verify_fire_incident(title, content, url, country="USA")
        fire_related_score = get_fire_related_score(content) if verification_result.lower().startswith("yes") else 0
        verified_at = datetime.now().isoformat()
        print(f"{date} {verification_result.strip().lower()} {url}")
        if verification_result.lower().startswith("yes"):
            row = {
                "title": title,
                "content": content,
                "published_date": date,
                "url": url,
                "source": source,
                "fire_related_score": fire_related_score,
                "verification_result": verification_result,
                "verified_at": verified_at
            }
            verified_rows.append(row)
            # Live update JSON for every 'yes' entry
            update_live_json(live_json_path, row)
            # Save to Excel immediately (append mode)
            if os.path.exists(excel_path):
                df_existing = pd.read_excel(excel_path)
                df_new = pd.DataFrame([row])
                df_combined = pd.concat([df_existing, df_new], ignore_index=True)
                df_combined.to_excel(excel_path, index=False)
            else:
                df = pd.DataFrame([row])
                df.to_excel(excel_path, index=False)
            autosize_and_format_excel(excel_path)
    if verified_rows:
        print(f"✅ Saved {len(verified_rows)} verified fire incidents to {excel_path} and {live_json_path}")
        # --- EMAIL SENDING LOGIC ---
        import smtplib
        from email.message import EmailMessage
        EMAIL_HOST = 'smtp.gmail.com'
        EMAIL_PORT = 587
        EMAIL_HOST_USER = 'agilemorphsolutions@gmail.com'
        EMAIL_HOST_PASSWORD = 'vktnzpaaurneigpg'
        TO_EMAIL = 'forrohitsingh99@gmail.com'
        def send_email_with_attachment(subject, body, to_email, attachment_paths):
            msg = EmailMessage()
            msg['Subject'] = subject
            msg['From'] = EMAIL_HOST_USER
            msg['To'] = to_email
            msg.set_content(body)
            for attachment_path in attachment_paths:
                with open(attachment_path, 'rb') as f:
                    file_data = f.read()
                    file_name = os.path.basename(attachment_path)
                msg.add_attachment(file_data, maintype='application', subtype='octet-stream', filename=file_name)
            with smtplib.SMTP(EMAIL_HOST, EMAIL_PORT) as server:
                server.starttls()
                server.login(EMAIL_HOST_USER, EMAIL_HOST_PASSWORD)
                server.send_message(msg)
        # Send the email with both the Excel and JSON files
        send_email_with_attachment(
            subject="Verified Fire Incidents - Latest Batch",
            body="Please find attached the latest verified fire incidents (Excel and JSON).",
            to_email=TO_EMAIL,
            attachment_paths=[excel_path, live_json_path]
        )
        print(f"📧 Email sent to {TO_EMAIL} with attachments {excel_path} and {live_json_path}")
    else:
        print("No verified fire incidents found.")

if __name__ == "__main__":
    import sys
    # Default to 'tweets_cleaned.json' if no argument is given
    json_path = sys.argv[1] if len(sys.argv) > 1 else "tweets_cleaned.json"
    excel_path = "output/verified_fires.xlsx"
    live_json_path = "output/live_verified_fires.json"
    # If input is tweets_raw.json, create tweets_cleaned.json first
    if os.path.basename(json_path) == "tweets_raw.json":
        cleaned_path = "tweets_cleaned.json"
        clean_tweets_json(json_path, cleaned_path)
        json_path = cleaned_path
    verify_and_save_to_excel(json_path, excel_path, live_json_path) 